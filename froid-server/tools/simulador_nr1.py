"""Gera a planilha de simulacao comercial do FROID NR-1.

Tudo que depende de uma entrada e FORMULA do Excel, nao numero calculado aqui.
A diferenca importa: o Fabio precisa mexer no efetivo, na base e nas faixas e
ver o resultado mudar. Uma planilha com valores fixos seria um relatorio.

Alem da receita, a planilha responde a pergunta que o preco nao responde: com
quantas respostas cada recorte por setor sai publicado. Os pisos vem do banco
(migrations/010 e 023) e estao replicados aqui como constantes documentadas —
se mudarem la, mudam aqui.

Uso:  python tools/simulador_nr1.py [caminho_de_saida.xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Pisos de anonimato, espelhados do banco.
# froid_nr1_min_cohort_total()      -> migrations/010_nr1_psychosocial_compliance.sql
# froid_nr1_min_cohort_cut()        -> migrations/010_nr1_psychosocial_compliance.sql
# froid_nr1_min_response_coverage() -> migrations/023_response_completeness.sql
PISO_CAMPANHA = 50
PISO_RECORTE = 10
COBERTURA_MINIMA = 0.50

AZUL = "1E293B"
AZUL_CLARO = "334155"
AMBAR = "451A03"
AMBAR_TEXTO = "FCD34D"
ENTRADA = "0C4A6E"
BRANCO = "F8FAFC"
CINZA = "94A3B8"

fina = Side(style="thin", color=AZUL_CLARO)
BORDA = Border(left=fina, right=fina, top=fina, bottom=fina)

MOEDA = 'R$ #,##0.00'
MOEDA0 = 'R$ #,##0'
PCT = '0.0%'


def _titulo(ws, linha: int, texto: str, largura: int = 6) -> None:
    ws.cell(row=linha, column=1, value=texto)
    c = ws.cell(row=linha, column=1)
    c.font = Font(bold=True, size=12, color=BRANCO)
    c.fill = PatternFill("solid", fgColor=AZUL)
    for col in range(2, largura + 1):
        ws.cell(row=linha, column=col).fill = PatternFill("solid", fgColor=AZUL)


def _rotulo(ws, linha: int, col: int, texto: str, nota: str = "") -> None:
    c = ws.cell(row=linha, column=col, value=texto)
    c.font = Font(bold=True)
    c.border = BORDA
    if nota:
        n = ws.cell(row=linha, column=col + 2, value=nota)
        n.font = Font(italic=True, size=9, color=CINZA)


def _entrada(ws, linha: int, col: int, valor, formato: str | None = None) -> None:
    c = ws.cell(row=linha, column=col, value=valor)
    c.fill = PatternFill("solid", fgColor=ENTRADA)
    c.font = Font(bold=True, color=BRANCO)
    c.border = BORDA
    if formato:
        c.number_format = formato


def _formula(ws, linha: int, col: int, f: str, formato: str | None = None,
             destaque: bool = False) -> None:
    c = ws.cell(row=linha, column=col, value=f)
    c.border = BORDA
    if formato:
        c.number_format = formato
    if destaque:
        c.fill = PatternFill("solid", fgColor=AMBAR)
        c.font = Font(bold=True, color=AMBAR_TEXTO)


def aba_receita(wb: Workbook):
    ws = wb.create_sheet("Receita")
    ws.sheet_view.showGridLines = False

    _titulo(ws, 1, "  FROID NR-1 — simulador de receita", 7)
    ws.cell(row=2, column=1,
            value="Células escuras são entradas: mude e o resto recalcula. "
                  "As faixas são cumulativas e incidem sobre o efetivo total.").font = Font(
        italic=True, size=9, color=CINZA)

    _titulo(ws, 4, "  Cliente", 7)
    _rotulo(ws, 5, 1, "Trabalhadores (efetivo total)")
    _entrada(ws, 5, 2, 300)
    _rotulo(ws, 6, 1, "Unidades (estabelecimentos)",
            "departamento no mesmo endereço NÃO é unidade")
    _entrada(ws, 6, 2, 1)
    _rotulo(ws, 7, 1, "Desconto aplicado")
    _entrada(ws, 7, 2, 0.15, PCT)

    _titulo(ws, 9, "  Tabela de preços", 7)
    ws.cell(row=10, column=1, value="Componente").font = Font(bold=True, color=BRANCO)
    ws.cell(row=10, column=2, value="Valor").font = Font(bold=True, color=BRANCO)
    ws.cell(row=10, column=3, value="Até").font = Font(bold=True, color=BRANCO)
    for col in (1, 2, 3):
        ws.cell(row=10, column=col).fill = PatternFill("solid", fgColor=AZUL_CLARO)

    _rotulo(ws, 11, 1, "Base por unidade / mês")
    _entrada(ws, 11, 2, 1200, MOEDA)
    _rotulo(ws, 12, 1, "Faixa 1 — por trabalhador / mês")
    _entrada(ws, 12, 2, 9.0, MOEDA)
    _entrada(ws, 12, 3, 100)
    _rotulo(ws, 13, 1, "Faixa 2 — por trabalhador / mês")
    _entrada(ws, 13, 2, 7.0, MOEDA)
    _entrada(ws, 13, 3, 300)
    _rotulo(ws, 14, 1, "Faixa 3 — por trabalhador / mês")
    _entrada(ws, 14, 2, 5.0, MOEDA)
    _entrada(ws, 14, 3, 1000)
    _rotulo(ws, 15, 1, "Faixa 4 — por trabalhador / mês")
    _entrada(ws, 15, 2, 3.0, MOEDA)
    ws.cell(row=15, column=3, value="acima").font = Font(italic=True, color=CINZA)

    _rotulo(ws, 17, 1, "Implantação (cobrança única)",
            "hoje em branco na proposta")
    _entrada(ws, 17, 2, 0, MOEDA)

    _titulo(ws, 19, "  Decomposição da fatura mensal", 7)
    ws.cell(row=20, column=1, value="Componente").font = Font(bold=True)
    ws.cell(row=20, column=2, value="Trabalhadores na faixa").font = Font(bold=True)
    ws.cell(row=20, column=3, value="Subtotal / mês").font = Font(bold=True)
    for col in (1, 2, 3):
        ws.cell(row=20, column=col).fill = PatternFill("solid", fgColor=AZUL_CLARO)
        ws.cell(row=20, column=col).font = Font(bold=True, color=BRANCO)
        ws.cell(row=20, column=col).alignment = Alignment(wrap_text=True)

    _rotulo(ws, 21, 1, "Base da plataforma")
    _formula(ws, 21, 2, "=B6", "0")
    _formula(ws, 21, 3, "=B6*B11", MOEDA)

    # MEDIAN recorta o efetivo dentro de cada faixa sem aninhar SE: devolve o
    # valor do meio entre zero, o efetivo e o teto da faixa.
    _rotulo(ws, 22, 1, "Faixa 1 (1 a 100)")
    _formula(ws, 22, 2, "=MEDIAN(0,$B$5,$C$12)", "0")
    _formula(ws, 22, 3, "=B22*B12", MOEDA)

    _rotulo(ws, 23, 1, "Faixa 2 (101 a 300)")
    _formula(ws, 23, 2, "=MEDIAN(0,$B$5-$C$12,$C$13-$C$12)", "0")
    _formula(ws, 23, 3, "=B23*B13", MOEDA)

    _rotulo(ws, 24, 1, "Faixa 3 (301 a 1.000)")
    _formula(ws, 24, 2, "=MEDIAN(0,$B$5-$C$13,$C$14-$C$13)", "0")
    _formula(ws, 24, 3, "=B24*B14", MOEDA)

    _rotulo(ws, 25, 1, "Faixa 4 (acima de 1.000)")
    _formula(ws, 25, 2, "=MAX(0,$B$5-$C$14)", "0")
    _formula(ws, 25, 3, "=B25*B15", MOEDA)

    _rotulo(ws, 26, 1, "Confere: soma das faixas")
    _formula(ws, 26, 2, "=SUM(B22:B25)", "0")
    ws.cell(row=26, column=3,
            value="deve igualar o efetivo").font = Font(italic=True, size=9, color=CINZA)

    _titulo(ws, 28, "  Resultado", 7)
    _rotulo(ws, 29, 1, "Mensal — tabela cheia")
    _formula(ws, 29, 2, "=SUM(C21:C25)", MOEDA)
    _rotulo(ws, 30, 1, "Mensal — com desconto")
    _formula(ws, 30, 2, "=B29*(1-B7)", MOEDA, destaque=True)
    _rotulo(ws, 31, 1, "Anual — com desconto")
    _formula(ws, 31, 2, "=B30*12", MOEDA0, destaque=True)
    _rotulo(ws, 32, 1, "Ano 1 — anual + implantação")
    _formula(ws, 32, 2, "=B31+B17", MOEDA0, destaque=True)
    _rotulo(ws, 33, 1, "Por trabalhador / ano")
    _formula(ws, 33, 2, '=IF(B5=0,"",B31/B5)', MOEDA)
    _rotulo(ws, 34, 1, "Por trabalhador / mês")
    _formula(ws, 34, 2, '=IF(B5=0,"",B30/B5)', MOEDA)
    _rotulo(ws, 35, 1, "Campanhas por ano")
    _entrada(ws, 35, 2, 2)
    ws.cell(row=35, column=4,
            value="linha de base + verificação de eficácia").font = Font(
        italic=True, size=9, color=CINZA)
    _rotulo(ws, 36, 1, "Receita por campanha")
    _formula(ws, 36, 2, '=IF(B35=0,"",B31/B35)', MOEDA0)
    ws.cell(row=37, column=1,
            value="É esse número que o cliente compara com um orçamento de "
                  "consultoria. Vale saber qual é antes da reunião.").font = Font(
        italic=True, size=9, color=CINZA)

    for col, larg in ((1, 34), (2, 18), (3, 18), (4, 34), (5, 14), (6, 14), (7, 14)):
        ws.column_dimensions[get_column_letter(col)].width = larg
    return ws


def aba_cenarios(wb: Workbook):
    ws = wb.create_sheet("Cenários")
    ws.sheet_view.showGridLines = False

    _titulo(ws, 1, "  Mesma tabela, efetivos diferentes", 6)
    ws.cell(row=2, column=1,
            value="Usa a tabela de preços da aba Receita. Mude lá e esta recalcula. "
                  "Unidades fixadas em 1.").font = Font(italic=True, size=9, color=CINZA)

    cabec = ["Trabalhadores", "Mensal", "Anual", "Por trab./ano",
             "Por trab./mês", "Receita/campanha"]
    for i, t in enumerate(cabec, start=1):
        c = ws.cell(row=4, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    efetivos = [50, 75, 100, 150, 200, 300, 500, 750, 1000, 1500, 3000, 5000]
    for i, n in enumerate(efetivos):
        L = 5 + i
        _entrada(ws, L, 1, n)
        mensal = (
            "=Receita!$B$11"
            "+MEDIAN(0,A{L},Receita!$C$12)*Receita!$B$12"
            "+MEDIAN(0,A{L}-Receita!$C$12,Receita!$C$13-Receita!$C$12)*Receita!$B$13"
            "+MEDIAN(0,A{L}-Receita!$C$13,Receita!$C$14-Receita!$C$13)*Receita!$B$14"
            "+MAX(0,A{L}-Receita!$C$14)*Receita!$B$15"
        ).format(L=L)
        _formula(ws, L, 2, mensal, MOEDA0)
        _formula(ws, L, 3, f"=B{L}*12", MOEDA0)
        _formula(ws, L, 4, f"=C{L}/A{L}", MOEDA)
        _formula(ws, L, 5, f"=B{L}/A{L}", MOEDA)
        _formula(ws, L, 6, f'=IF(Receita!$B$35=0,"",C{L}/Receita!$B$35)', MOEDA0)

    aviso = ws.cell(
        row=5 + len(efetivos) + 1, column=1,
        value="Os efetivos de 50 e 75 estão aqui de propósito: abaixo de cerca de "
              "75 a campanha não atinge o piso de anonimato e não produz resultado "
              "liberável. Ver a aba Anonimato antes de propor.")
    aviso.font = Font(italic=True, size=9, color=AMBAR_TEXTO)
    aviso.fill = PatternFill("solid", fgColor=AMBAR)

    for col, larg in ((1, 16), (2, 16), (3, 16), (4, 16), (5, 16), (6, 18)):
        ws.column_dimensions[get_column_letter(col)].width = larg
    return ws


def aba_anonimato(wb: Workbook):
    ws = wb.create_sheet("Anonimato")
    ws.sheet_view.showGridLines = False

    _titulo(ws, 1, "  Quantas respostas para o resultado sair", 6)
    ws.cell(row=2, column=1,
            value="Os pisos NÃO são configuráveis comercialmente: estão no banco e "
                  "existem para proteger o trabalhador. Esta aba diz o que a adesão "
                  "precisa ser, não o contrário.").font = Font(
        italic=True, size=9, color=CINZA)

    _titulo(ws, 4, "  Pisos do sistema", 6)
    _rotulo(ws, 5, 1, "Piso da campanha (respostas substantivas)")
    ws.cell(row=5, column=2, value=PISO_CAMPANHA).font = Font(bold=True)
    ws.cell(row=5, column=4,
            value="froid_nr1_min_cohort_total()").font = Font(italic=True, size=9, color=CINZA)
    _rotulo(ws, 6, 1, "Piso por recorte (setor / unidade)")
    ws.cell(row=6, column=2, value=PISO_RECORTE).font = Font(bold=True)
    ws.cell(row=6, column=4,
            value="froid_nr1_min_cohort_cut()").font = Font(italic=True, size=9, color=CINZA)
    _rotulo(ws, 7, 1, "Cobertura mínima para a resposta contar")
    ws.cell(row=7, column=2, value=COBERTURA_MINIMA).number_format = PCT
    ws.cell(row=7, column=2).font = Font(bold=True)
    ws.cell(row=7, column=4,
            value="das dimensões do instrumento").font = Font(italic=True, size=9, color=CINZA)

    _titulo(ws, 9, "  Adesão esperada", 6)
    _rotulo(ws, 10, 1, "Adesão (respondem)")
    _entrada(ws, 10, 2, 0.70, PCT)
    _rotulo(ws, 11, 1, "Dos que respondem, quantos completam o suficiente")
    _entrada(ws, 11, 2, 0.85, PCT)
    _rotulo(ws, 12, 1, "Taxa efetiva (substantivas / convidados)")
    _formula(ws, 12, 2, "=B10*B11", PCT, destaque=True)

    _titulo(ws, 14, "  Departamentos", 6)
    cabec = ["Departamento", "Efetivo", "Respostas substantivas previstas",
             "Publica?", "Adesão mínima para publicar", "Folga"]
    for i, t in enumerate(cabec, start=1):
        c = ws.cell(row=15, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    exemplo = [("Operações", 120), ("Comercial", 80), ("Administrativo", 50),
               ("Logística", 30), ("Diretoria", 20)]
    primeira, ultima = 16, 16 + len(exemplo) - 1
    for i, (nome, efetivo) in enumerate(exemplo):
        L = primeira + i
        _entrada(ws, L, 1, nome)
        _entrada(ws, L, 2, efetivo)
        _formula(ws, L, 3, f"=ROUND(B{L}*$B$12,0)", "0")
        _formula(ws, L, 4, f'=IF(C{L}>=$B$6,"sim","SUPRIMIDO")')
        _formula(ws, L, 5, f'=IF(B{L}=0,"",$B$6/B{L}/$B$11)', PCT)
        _formula(ws, L, 6, f"=C{L}-$B$6", "0")

    total = ultima + 1
    _rotulo(ws, total, 1, "Total da campanha")
    _formula(ws, total, 2, f"=SUM(B{primeira}:B{ultima})", "0")
    _formula(ws, total, 3, f"=SUM(C{primeira}:C{ultima})", "0")
    _formula(ws, total, 4,
             f'=IF(C{total}>=$B$5,"campanha liberada","ABAIXO DO PISO")',
             destaque=True)

    ws.cell(row=total + 2, column=1,
            value="O menor departamento é quem manda. Um setor de 20 pessoas precisa "
                  "de metade dele respondendo para existir no relatório — e é sobre "
                  "esse setor que o cliente vai perguntar na reunião.").font = Font(
        italic=True, size=9, color=AMBAR_TEXTO)

    ws.cell(row=total + 4, column=1,
            value="Recorte suprimido não é falha do FROID: é o piso de anonimato "
                  "funcionando. Combinar isso ANTES do contrato evita a conversa "
                  "ruim depois da campanha encerrada.").font = Font(
        italic=True, size=9, color=CINZA)

    for col, larg in ((1, 30), (2, 14), (3, 22), (4, 20), (5, 22), (6, 12)):
        ws.column_dimensions[get_column_letter(col)].width = larg
    return ws


def aba_alavancas(wb: Workbook):
    ws = wb.create_sheet("Alavancas")
    ws.sheet_view.showGridLines = False

    _titulo(ws, 1, "  E se a tabela mudasse?", 7)
    ws.cell(row=2, column=1,
            value="Compara a tabela vigente com cenários de reajuste, no efetivo "
                  "informado na aba Receita. Nada aqui altera a tabela real.").font = Font(
        italic=True, size=9, color=CINZA)

    _rotulo(ws, 4, 1, "Efetivo simulado")
    _formula(ws, 4, 2, "=Receita!B5", "0")

    cabec = ["Cenário", "Base/mês", "Faixa 1", "Faixa 2", "Faixa 3", "Faixa 4",
             "Mensal", "Anual", "Variação"]
    for i, t in enumerate(cabec, start=1):
        c = ws.cell(row=6, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    cenarios = [
        ("Vigente", 1200, 9.0, 7.0, 5.0, 3.0),
        ("Base R$ 2.500", 2500, 9.0, 7.0, 5.0, 3.0),
        ("Base R$ 3.500", 3500, 9.0, 7.0, 5.0, 3.0),
        ("Faixas +25%", 1200, 11.25, 8.75, 6.25, 3.75),
        ("Faixas +40%", 1200, 12.60, 9.80, 7.00, 4.20),
        ("Base 2.500 + faixas +25%", 2500, 11.25, 8.75, 6.25, 3.75),
    ]
    primeira = 7
    for i, (nome, base, f1, f2, f3, f4) in enumerate(cenarios):
        L = primeira + i
        ws.cell(row=L, column=1, value=nome).font = Font(bold=True)
        ws.cell(row=L, column=1).border = BORDA
        for col, val, fmt in ((2, base, MOEDA0), (3, f1, MOEDA), (4, f2, MOEDA),
                              (5, f3, MOEDA), (6, f4, MOEDA)):
            _entrada(ws, L, col, val, fmt)
        mensal = (
            "=B{L}"
            "+MEDIAN(0,$B$4,Receita!$C$12)*C{L}"
            "+MEDIAN(0,$B$4-Receita!$C$12,Receita!$C$13-Receita!$C$12)*D{L}"
            "+MEDIAN(0,$B$4-Receita!$C$13,Receita!$C$14-Receita!$C$13)*E{L}"
            "+MAX(0,$B$4-Receita!$C$14)*F{L}"
        ).format(L=L)
        _formula(ws, L, 7, mensal, MOEDA0, destaque=(i == 0))
        _formula(ws, L, 8, f"=G{L}*12", MOEDA0)
        _formula(ws, L, 9, f'=IF($G${primeira}=0,"",G{L}/$G${primeira}-1)', PCT)

    ws.cell(row=primeira + len(cenarios) + 2, column=1,
            value="A implantação não está aqui de propósito: ela é cobrança única "
                  "e não muda a recorrência. Some na aba Receita, célula B17.").font = Font(
        italic=True, size=9, color=CINZA)

    for col, larg in ((1, 26), (2, 14), (3, 12), (4, 12), (5, 12), (6, 12),
                      (7, 16), (8, 16), (9, 12)):
        ws.column_dimensions[get_column_letter(col)].width = larg
    return ws


def aba_concorrencia(wb: Workbook):
    """Cotacoes de concorrentes, para preencher a mao.

    Sai VAZIA de proposito. Preencher com numero inventado seria pior do que
    nao ter aba nenhuma: viraria referencia numa reuniao e ninguem lembraria
    que a origem era um chute. As formulas ja estao prontas e comecam a
    devolver resultado assim que houver dado real.
    """
    ws = wb.create_sheet("Concorrência")
    ws.sheet_view.showGridLines = False

    _titulo(ws, 1, "  Concorrência — preencher com cotações reais", 9)
    ws.cell(row=2, column=1,
            value="Nada aqui foi estimado. Preencha as células escuras com valores "
                  "de propostas, tabelas públicas ou cotações recebidas — e anote a "
                  "fonte e a data, porque preço sem procedência não sustenta "
                  "argumento em reunião.").font = Font(italic=True, size=9, color=CINZA)

    _titulo(ws, 4, "  Referência FROID (vem da aba Receita)", 9)
    _rotulo(ws, 5, 1, "Efetivo simulado")
    _formula(ws, 5, 2, "=Receita!B5", "0")
    _rotulo(ws, 6, 1, "Anual FROID (com desconto)")
    _formula(ws, 6, 2, "=Receita!B31", MOEDA0)
    _rotulo(ws, 7, 1, "Implantação FROID")
    _formula(ws, 7, 2, "=Receita!B17", MOEDA0)
    _rotulo(ws, 8, 1, "Ano 1 FROID (anual + implantação)")
    _formula(ws, 8, 2, "=Receita!B32", MOEDA0, destaque=True)
    _rotulo(ws, 9, 1, "Campanhas/ano FROID")
    _formula(ws, 9, 2, "=Receita!B35", "0")

    _titulo(ws, 10, "  Cotações", 9)
    nota = ws.cell(
        row=11, column=1,
        value="Preencha “Mensal recorrente” OU “Por campanha”, conforme o modelo da "
              "cotação. Se os dois estiverem preenchidos, o cálculo usa o mensal. "
              "“Campanhas/ano” em branco assume o mesmo número do FROID, para que a "
              "comparação seja sobre o mesmo ciclo.")
    nota.font = Font(italic=True, size=9, color=CINZA)
    cabec = [
        "Concorrente", "Modelo de cobrança", "Efetivo da cotação",
        "Implantação (única)", "Mensal recorrente", "Por campanha",
        "Campanhas/ano incluídas", "Prazo de entrega (dias)",
        "Fonte e data da informação",
    ]
    for i, t in enumerate(cabec, start=1):
        c = ws.cell(row=12, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 42

    primeira, n_linhas = 13, 6
    ultima = primeira + n_linhas - 1
    for L in range(primeira, ultima + 1):
        for col in range(1, 10):
            fmt = MOEDA0 if col in (4, 5, 6) else ("0" if col in (3, 7, 8) else None)
            _entrada(ws, L, col, None, fmt)

    _titulo(ws, ultima + 2, "  Comparação — calculada a partir do que for preenchido", 9)
    cab2 = [
        "Concorrente", "Anual do concorrente", "Ano 1 do concorrente",
        "Por trabalhador / ano", "Por campanha", "Ano 1 vs FROID (R$)",
        "Ano 1 vs FROID (%)", "FROID está…", "",
    ]
    L0 = ultima + 3
    for i, t in enumerate(cab2, start=1):
        c = ws.cell(row=L0, column=i, value=t)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[L0].height = 42

    for i in range(n_linhas):
        origem = primeira + i
        L = L0 + 1 + i
        _formula(ws, L, 1, f'=IF(A{origem}="","",A{origem})')
        # Anual = 12 mensalidades OU campanhas x valor por campanha, conforme o
        # modelo que a cotacao usar. Aceita os dois sem obrigar a escolher.
        _formula(
            ws, L, 2,
            f'=IF(AND(E{origem}="",F{origem}=""),"",'
            f'IF(E{origem}<>"",E{origem}*12,F{origem}*IF(G{origem}="",$B$9,G{origem})))',
            MOEDA0)
        _formula(ws, L, 3, f'=IF(B{L}="","",B{L}+IF(D{origem}="",0,D{origem}))', MOEDA0)
        _formula(
            ws, L, 4,
            f'=IF(OR(B{L}="",C{origem}="",C{origem}=0),"",B{L}/C{origem})', MOEDA)
        _formula(
            ws, L, 5,
            f'=IF(B{L}="","",B{L}/IF(G{origem}="",$B$9,G{origem}))', MOEDA0)
        _formula(ws, L, 6, f'=IF(C{L}="","",$B$8-C{L})', MOEDA0)
        _formula(ws, L, 7, f'=IF(OR(C{L}="",C{L}=0),"",$B$8/C{L}-1)', PCT)
        _formula(
            ws, L, 8,
            f'=IF(G{L}="","",IF(G{L}>0.05,"mais caro",'
            f'IF(G{L}<-0.05,"mais barato","equivalente")))', destaque=True)

    aviso = ws.cell(
        row=L0 + n_linhas + 2, column=1,
        value="Cuidado ao comparar: quase nenhum concorrente entrega prova de "
              "eficácia antes/depois nem piso de anonimato. Preço menor por um "
              "escopo menor não é preço menor — é outro produto. A matriz abaixo "
              "existe para não deixar isso passar.")
    aviso.font = Font(italic=True, size=9, color=AMBAR_TEXTO)
    aviso.fill = PatternFill("solid", fgColor=AMBAR)
    ws.row_dimensions[L0 + n_linhas + 2].height = 32

    L1 = L0 + n_linhas + 4
    _titulo(ws, L1, "  O que cada um entrega (preencher: sim / não / parcial)", 9)
    # Coluna 1 = capacidade, coluna 2 = FROID, colunas 3..8 = concorrentes.
    # A linha de baixo puxa o nome digitado na tabela de cotacoes, para nao
    # obrigar a redigitar e nao deixar as duas metades da aba divergirem.
    for col, rotulo in [(1, "Capacidade"), (2, "FROID")]:
        c = ws.cell(row=L1 + 1, column=col, value=rotulo)
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(horizontal="center")
    for i in range(n_linhas):
        c = ws.cell(row=L1 + 1, column=3 + i, value=f"Concorrente {i + 1}")
        c.font = Font(bold=True, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        _formula(ws, L1 + 2, 3 + i, f'=IF(A{primeira + i}="","",A{primeira + i})')
    ws.cell(row=L1 + 2, column=1, value="(nome da cotação)").font = Font(
        italic=True, size=9, color=CINZA)

    capacidades = [
        ("Inventário com os nove campos (1.5.7.3.2)", "sim"),
        ("AEP psicossocial", "sim"),
        ("Gradação por severidade × probabilidade (1.5.4.4.4)", "sim"),
        ("Probabilidade desconta eficácia medida (1.5.4.4.5.3)", "sim"),
        ("Prova de eficácia antes/depois, setor contra si mesmo", "sim"),
        ("Medida ineficaz não reduz risco no papel (1.5.5.3.2.1)", "sim"),
        ("Piso de anonimato por recorte", "sim"),
        ("Resultado só após encerrar a campanha", "sim"),
        ("Recusa abrir campanha sem canal de apoio", "sim"),
        ("Critérios de gradação alinhados ao PGR da empresa", "sim"),
        ("Guarda do histórico por 20 anos (1.5.7.3.3.1)", "sim"),
        ("Painel próprio, sem planilha intermediária", "sim"),
        ("Base legal declarada (obrigação legal, não consentimento)", "sim"),
        ("Não entrega resposta individual a ninguém", "sim"),
    ]
    for j, (nome, froid) in enumerate(capacidades):
        L = L1 + 3 + j
        c = ws.cell(row=L, column=1, value=nome)
        c.font = Font(size=10)
        c.border = BORDA
        c.alignment = Alignment(wrap_text=True)
        v = ws.cell(row=L, column=2, value=froid)
        v.font = Font(bold=True, color=AMBAR_TEXTO)
        v.fill = PatternFill("solid", fgColor=AMBAR)
        v.alignment = Alignment(horizontal="center")
        v.border = BORDA
        for i in range(n_linhas):
            _entrada(ws, L, 3 + i, None)

    ws.cell(row=L1 + 3 + len(capacidades) + 1, column=1,
            value="A coluna FROID reflete o que está implementado no produto hoje. "
                  "Se alguma linha deixar de ser verdade, corrija aqui antes de usar "
                  "a matriz com um cliente.").font = Font(italic=True, size=9, color=CINZA)

    larguras = [(1, 46), (2, 22), (3, 18), (4, 18), (5, 18), (6, 16), (7, 18),
                (8, 18), (9, 30)]
    for col, larg in larguras:
        ws.column_dimensions[get_column_letter(col)].width = larg
    return ws


def aba_leia(wb: Workbook):
    ws = wb.active
    ws.title = "Leia primeiro"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 108

    linhas = [
        ("t", "FROID NR-1 — simulador comercial"),
        ("", ""),
        ("n", "Células de fundo escuro são ENTRADAS. Todo o resto é fórmula e "
              "recalcula sozinho."),
        ("", ""),
        ("h", "Receita"),
        ("n", "A conta de uma proposta. Efetivo, unidades, desconto e a tabela de "
              "preços de um lado; fatura decomposta faixa a faixa do outro."),
        ("n", "Unidade é ESTABELECIMENTO, não departamento. Cinco setores no mesmo "
              "endereço são uma base, não cinco. Errar isso multiplica a fatura."),
        ("", ""),
        ("h", "Cenários"),
        ("n", "A mesma tabela aplicada a doze efetivos. Serve para ver onde o preço "
              "por trabalhador cai e onde o contrato fica pequeno demais para o "
              "trabalho que dá."),
        ("", ""),
        ("h", "Anonimato"),
        ("n", "A parte que preço nenhum resolve. Diz quantas respostas cada setor "
              "precisa para aparecer no relatório."),
        ("n", "Os pisos vêm do banco e não são negociáveis comercialmente. Um "
              "recorte suprimido é o piso funcionando, não defeito — mas isso "
              "precisa estar combinado ANTES do contrato."),
        ("", ""),
        ("h", "Alavancas"),
        ("n", "Compara a tabela vigente com cenários de reajuste. Não altera nada: "
              "é só para decidir."),
        ("", ""),
        ("h", "Concorrência"),
        ("n", "Sai VAZIA de propósito — nenhum valor foi estimado. Preencha com "
              "cotações reais e anote fonte e data: preço sem procedência não "
              "sustenta argumento em reunião."),
        ("n", "Aceita cobrança mensal ou por campanha, e calcula sozinha anual, "
              "ano 1, custo por trabalhador e a diferença contra o FROID."),
        ("n", "Embaixo há a matriz de capacidades. Ela existe porque preço menor "
              "por escopo menor não é preço menor — é outro produto, e a "
              "diferença precisa aparecer na mesma tela."),
        ("", ""),
        ("a", "O que esta planilha NÃO faz: não estima custo, não tem margem e não "
              "conhece preço de concorrente. É receita bruta. Enquanto o custo de "
              "operar uma campanha não estiver medido, ela mostra quanto entra — "
              "não quanto sobra."),
    ]
    L = 2
    for tipo, texto in linhas:
        c = ws.cell(row=L, column=1, value=texto)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if tipo == "t":
            c.font = Font(bold=True, size=15)
        elif tipo == "h":
            c.font = Font(bold=True, size=11, color=BRANCO)
            c.fill = PatternFill("solid", fgColor=AZUL)
        elif tipo == "a":
            c.font = Font(italic=True, color=AMBAR_TEXTO)
            c.fill = PatternFill("solid", fgColor=AMBAR)
            ws.row_dimensions[L].height = 58
        else:
            c.font = Font(size=10)
            ws.row_dimensions[L].height = 30
        L += 1
    return ws


def main() -> int:
    destino = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("FROID_NR1_simulador.xlsx")
    wb = Workbook()
    aba_leia(wb)
    aba_receita(wb)
    aba_cenarios(wb)
    aba_anonimato(wb)
    aba_alavancas(wb)
    aba_concorrencia(wb)
    wb.save(destino)
    print(f"planilha gerada: {destino.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
